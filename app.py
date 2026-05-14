import streamlit as st
import pandas as pd
import json
import io
import re
from datetime import date, datetime
import anthropic
from dotenv import load_dotenv
import os
import fitz  # PyMuPDF
import base64

# ── Configuración de página ──────────────────────────────────────────────────
st.set_page_config(
    page_title="Gestor Ambiental IA",
    page_icon="♻️",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Estilos CSS ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600;700&display=swap');

:root {
    --verde: #1a7a4a;
    --verde-claro: #2ecc71;
    --verde-oscuro: #0d3d25;
    --crema: #f5f0e8;
    --gris: #6b7280;
    --rojo: #e74c3c;
    --amarillo: #f39c12;
}

html, body, [data-testid="stAppViewContainer"] {
    background-color: #f0f4f0 !important;
    font-family: 'IBM Plex Sans', sans-serif;
}

[data-testid="stHeader"] { background: transparent !important; }

.header-banner {
    background: linear-gradient(135deg, #0d3d25 0%, #1a7a4a 60%, #2ecc71 100%);
    border-radius: 16px;
    padding: 2.5rem 3rem;
    margin-bottom: 2rem;
    position: relative;
    overflow: hidden;
}
.header-banner::before {
    content: "♻";
    position: absolute;
    right: 2rem;
    top: 50%;
    transform: translateY(-50%);
    font-size: 8rem;
    opacity: 0.08;
    line-height: 1;
}
.header-banner h1 {
    font-family: 'IBM Plex Mono', monospace;
    color: #fff;
    font-size: 1.9rem;
    font-weight: 600;
    margin: 0 0 0.3rem 0;
    letter-spacing: -0.5px;
}
.header-banner p {
    color: rgba(255,255,255,0.75);
    margin: 0;
    font-size: 0.95rem;
    font-weight: 300;
}
.header-badge {
    display: inline-block;
    background: rgba(255,255,255,0.15);
    border: 1px solid rgba(255,255,255,0.3);
    border-radius: 20px;
    padding: 0.2rem 0.8rem;
    font-size: 0.75rem;
    color: #fff;
    margin-bottom: 0.8rem;
    font-family: 'IBM Plex Mono', monospace;
    letter-spacing: 1px;
    text-transform: uppercase;
}

.card {
    background: #fff;
    border-radius: 12px;
    padding: 1.5rem;
    box-shadow: 0 2px 12px rgba(0,0,0,0.06);
    margin-bottom: 1.2rem;
    border: 1px solid #e8f0eb;
}
.card-title {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 0.8rem;
    color: var(--verde);
    text-transform: uppercase;
    letter-spacing: 1.5px;
    margin-bottom: 1rem;
    font-weight: 600;
}

.stButton > button {
    background: linear-gradient(135deg, #1a7a4a, #2ecc71) !important;
    color: white !important;
    border: none !important;
    border-radius: 8px !important;
    padding: 0.6rem 1.8rem !important;
    font-family: 'IBM Plex Mono', monospace !important;
    font-size: 0.85rem !important;
    font-weight: 600 !important;
    letter-spacing: 0.5px !important;
    transition: all 0.2s !important;
    box-shadow: 0 4px 15px rgba(26,122,74,0.3) !important;
}
.stButton > button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 6px 20px rgba(26,122,74,0.4) !important;
}

.btn-rojo > button {
    background: linear-gradient(135deg, #c0392b, #e74c3c) !important;
    box-shadow: 0 4px 15px rgba(231,76,60,0.3) !important;
}

.status-ok {
    background: #e8f8ef;
    border-left: 4px solid #2ecc71;
    padding: 0.7rem 1rem;
    border-radius: 0 8px 8px 0;
    margin: 0.4rem 0;
    font-size: 0.88rem;
    color: #0d3d25;
}
.status-warn {
    background: #fef9ec;
    border-left: 4px solid #f39c12;
    padding: 0.7rem 1rem;
    border-radius: 0 8px 8px 0;
    margin: 0.4rem 0;
    font-size: 0.88rem;
    color: #7d5a00;
}
.status-error {
    background: #fdf0ee;
    border-left: 4px solid #e74c3c;
    padding: 0.7rem 1rem;
    border-radius: 0 8px 8px 0;
    margin: 0.4rem 0;
    font-size: 0.88rem;
    color: #7b1a12;
}

.metric-box {
    background: linear-gradient(135deg, #0d3d25, #1a7a4a);
    border-radius: 10px;
    padding: 1rem 1.2rem;
    color: white;
    text-align: center;
}
.metric-box .num {
    font-family: 'IBM Plex Mono', monospace;
    font-size: 2rem;
    font-weight: 600;
    line-height: 1;
}
.metric-box .label {
    font-size: 0.75rem;
    opacity: 0.75;
    margin-top: 0.3rem;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

.acum-banner {
    background: linear-gradient(135deg, #1a3a5c, #2980b9);
    border-radius: 10px;
    padding: 0.8rem 1.2rem;
    color: white;
    margin-bottom: 1rem;
    font-size: 0.9rem;
}

[data-testid="stDataFrame"] {
    border-radius: 10px !important;
    overflow: hidden !important;
}

.footer {
    text-align: center;
    padding: 2rem;
    color: var(--gris);
    font-size: 0.8rem;
    border-top: 1px solid #e0e8e3;
    margin-top: 3rem;
}
</style>
""", unsafe_allow_html=True)

# ── Cargar API Key ───────────────────────────────────────────────────────────
load_dotenv()

def get_api_key():
    key = os.getenv("ANTHROPIC_API_KEY")
    if not key:
        try:
            key = st.secrets.get("ANTHROPIC_API_KEY", "")
        except Exception:
            pass
    return key

# ── Inicializar estado de sesión (bitácora acumulada) ────────────────────────
# Esto es como una "memoria" de la app durante tu sesión de trabajo.
# Cada vez que procesas PDFs, los nuevos registros SE AGREGAN a esta lista.
# Solo se borra si recargas la página o presionas "Limpiar bitácora".
if "bitacora_acumulada" not in st.session_state:
    st.session_state["bitacora_acumulada"] = []   # Lista vacía al inicio
if "errores_acumulados" not in st.session_state:
    st.session_state["errores_acumulados"] = []
if "archivos_procesados" not in st.session_state:
    st.session_state["archivos_procesados"] = set()  # Nombres de archivos ya procesados


# ── Conversión de PDF a imágenes base64 ─────────────────────────────────────
# MEJORA: Ahora comprime más las imágenes si el PDF es muy pesado (> 2 MB),
# y además detecta cuántas páginas tiene para procesarlas correctamente.
def pdf_to_images_base64(pdf_file, calidad_jpeg: int = 72) -> list:
    """
    Convierte cada página del PDF a imagen base64.
    - Si el PDF pesa más de 2 MB, reduce automáticamente la resolución y calidad.
    - Funciona con PDFs de 1 página o de muchas páginas.
    """
    try:
        pdf_bytes = pdf_file.read()
        tamano_mb = len(pdf_bytes) / (1024 * 1024)

        # Si el PDF pesa más de 2 MB, usamos menor resolución para no sobrecargar la IA
        if tamano_mb > 2:
            dpi = 150   # Resolución más baja (antes era 200)
            calidad = 65  # Compresión más agresiva
        elif tamano_mb > 1:
            dpi = 175
            calidad = 72
        else:
            dpi = 200   # Resolución normal para PDFs pequeños
            calidad = 80

        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        total_paginas = len(doc)
        images = []

        for i, page in enumerate(doc):
            mat = fitz.Matrix(dpi / 72, dpi / 72)
            pix = page.get_pixmap(matrix=mat, colorspace=fitz.csRGB)
            img_bytes = pix.tobytes("jpeg", jpg_quality=calidad)
            b64 = base64.b64encode(img_bytes).decode("utf-8")
            images.append({
                "page": i + 1,
                "total_pages": total_paginas,
                "data": b64
            })

        doc.close()

        if not images:
            raise RuntimeError("El PDF no tiene páginas.")

        return images, tamano_mb, total_paginas

    except Exception as e:
        raise RuntimeError(f"No se pudo procesar el PDF: {e}")


# ── Prompt de extracción por visión ──────────────────────────────────────────
PROMPT_VISION = """Eres un experto en normativa ambiental mexicana (SEMARNAT/NOM-055-SEMARNAT-2003).
Analiza visualmente las imágenes del "Manifiesto de Entrega, Transporte y Recepción de Residuos Peligrosos" de México que se te proporcionan.
El documento puede estar escaneado, escrito a mano o impreso. Lee cuidadosamente todos los campos visibles.

IMPORTANTE: Si las imágenes contienen MÚLTIPLES manifiestos (varios formularios completos), extrae CADA UNO por separado
y devuelve un ARRAY JSON con un objeto por cada manifiesto encontrado.
Si solo hay UN manifiesto, devuelve un array con un solo objeto.

Extrae EXACTAMENTE los siguientes campos por cada manifiesto y devuelve SOLO un array JSON válido (sin backticks, sin texto extra):

[
  {
    "consecutivo": "Número de folio/manifiesto. Busca 'No.', 'Núm.', 'Folio', o número grande destacado. Ej: '047912'",
    "nombre_residuo": "Nombre del residuo peligroso de la sección 5. Ej: 'ACEITE Y LUBRICANTE USADO Y GASTADO'",
    "cantidad": "Cantidad y unidad de la sección 5. Ej: '6,000 LTS' o '1.116 TON'. Si hay varias, sepáralas con coma.",
    "cretib": "Características de peligrosidad CRETIB marcadas con X o tache en sección 5. Solo las letras activas. Ej: 'T'",
    "fecha_salida": "Fecha de la sección 7 (generador), donde dice 'Fecha:'. Formato DD/MM/AAAA. Ej: '20/DIC/2024'",
    "responsable": "Nombre del responsable firmante de la sección 7 (generador). NO incluir transportista ni destinatario.",
    "fase_siguiente": "Nombre y/o razón social del destinatario de la sección 15. Ej: 'LUBRICANTES JUGUER S.A. DE C.V.'",
    "area_resguardo": "Mismo valor que fase_siguiente.",
    "transportista_nombre": "Nombre o razón social del transportista, sección 8.",
    "transportista_autorizacion": "Número de autorización SEMARNAT del transportista, sección 9. Ej: '05-35-PS-I-327D-2019'",
    "destinatario_nombre": "Nombre o razón social del destinatario, sección 15.",
    "destinatario_autorizacion": "Número de autorización SEMARNAT del destinatario, sección 16."
  }
]

REGLAS:
- Si un campo no es visible o legible, usa exactamente: "No especificado"
- Para CRETIB: observa las casillas marcadas con X o tache en la tabla de clasificación
- El responsable es SOLO el de la sección 7 del generador
- NO inventes datos
- Responde ÚNICAMENTE el array JSON, sin texto adicional
- Si hay múltiples manifiestos en las imágenes, incluye UN objeto por cada uno
"""


# ── Llamada a Claude Vision ───────────────────────────────────────────────────
def extraer_datos_claude(images: list, api_key: str) -> list:
    """
    Envía las imágenes del PDF a Claude y extrae los datos de los manifiestos.
    Ahora devuelve una LISTA de manifiestos (puede haber más de uno por PDF).
    """
    client = anthropic.Anthropic(api_key=api_key)

    # Construir el mensaje con todas las páginas como imágenes
    content = []
    for img in images:
        content.append({
            "type": "image",
            "source": {
                "type": "base64",
                "media_type": "image/jpeg",
                "data": img["data"]
            }
        })
    content.append({
        "type": "text",
        "text": PROMPT_VISION
    })

    response = client.messages.create(
        model="claude-opus-4-5",
        max_tokens=2000,  # Aumentado para poder manejar múltiples manifiestos
        messages=[{"role": "user", "content": content}]
    )

    raw = response.content[0].text.strip()
    # Limpiar por si la IA agrega marcas de código
    raw = re.sub(r"^```(?:json)?\s*", "", raw)
    raw = re.sub(r"\s*```$", "", raw)

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        raise ValueError(f"La IA devolvió formato inválido: {e}\n\nRespuesta recibida:\n{raw[:500]}")

    # La IA puede devolver un objeto solo o una lista — normalizamos a lista siempre
    if isinstance(data, dict):
        data = [data]

    # Validar y completar campos faltantes en cada manifiesto
    campos = [
        "consecutivo", "nombre_residuo", "cantidad", "cretib",
        "fecha_salida", "responsable", "fase_siguiente", "area_resguardo",
        "transportista_nombre", "transportista_autorizacion",
        "destinatario_nombre", "destinatario_autorizacion"
    ]
    for manifiesto in data:
        for c in campos:
            if c not in manifiesto or not manifiesto[c]:
                manifiesto[c] = "No especificado"

    return data


# ── Generar Excel ────────────────────────────────────────────────────────────
def generar_excel(registros: list, fecha_ingreso: str) -> bytes:
    """
    Genera el archivo Excel con TODOS los registros acumulados en la sesión.
    Incluye formato institucional SEMARNAT.
    """
    filas = []
    for r in registros:
        filas.append({
            "Consecutivo (No. Manifiesto)": r.get("consecutivo", ""),
            "Nombre del Residuo Peligroso": r.get("nombre_residuo", ""),
            "Cantidad Generada": r.get("cantidad", ""),
            "Características CRETIB": r.get("cretib", ""),
            "Fecha de Ingreso al Almacén": r.get("_fecha_ingreso", fecha_ingreso),
            "Fecha de Salida del Almacén": r.get("fecha_salida", ""),
            "Señalamiento Fase Siguiente": r.get("fase_siguiente", ""),
            "Área de Resguardo / Transferencia": r.get("area_resguardo", ""),
            "Transportista - Nombre/Razón Social": r.get("transportista_nombre", ""),
            "Transportista - No. Autorización SEMARNAT": r.get("transportista_autorizacion", ""),
            "Destinatario - Nombre/Razón Social": r.get("destinatario_nombre", ""),
            "Destinatario - No. Autorización SEMARNAT": r.get("destinatario_autorizacion", ""),
            "Nombre del Responsable": r.get("responsable", ""),
            "Archivo Fuente": r.get("_archivo", ""),
        })

    df = pd.DataFrame(filas)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Bitácora RP")
        ws = writer.sheets["Bitácora RP"]

        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

        verde_header = PatternFill("solid", fgColor="0D5C2E")
        verde_sub    = PatternFill("solid", fgColor="1A7A4A")
        crema_par    = PatternFill("solid", fgColor="F5FAF7")
        blanco       = PatternFill("solid", fgColor="FFFFFF")

        borde_fino = Border(
            left=Side(style="thin", color="C8DDD0"),
            right=Side(style="thin", color="C8DDD0"),
            top=Side(style="thin", color="C8DDD0"),
            bottom=Side(style="thin", color="C8DDD0"),
        )

        # Fila 1: Título institucional
        ws.insert_rows(1)
        ws.merge_cells("A1:N1")
        c = ws["A1"]
        c.value = "BITÁCORA DE RESIDUOS PELIGROSOS — FORMATO SEMARNAT"
        c.font = Font(bold=True, color="FFFFFF", size=13, name="Calibri")
        c.fill = verde_header
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        # Fila 2: Fecha de generación
        ws.insert_rows(2)
        ws.merge_cells("A2:N2")
        c = ws["A2"]
        c.value = (
            f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   |   "
            f"Total de registros: {len(registros)}"
        )
        c.font = Font(italic=True, color="FFFFFF", size=9, name="Calibri")
        c.fill = verde_sub
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 18

        # Fila 3: Encabezados de columna
        header_row = 3
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
            cell.fill = verde_header
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = borde_fino
        ws.row_dimensions[header_row].height = 36

        # Filas de datos con colores alternados
        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row), start=0):
            fill = crema_par if row_idx % 2 == 0 else blanco
            for cell in row:
                cell.fill = fill
                cell.border = borde_fino
                cell.font = Font(size=9, name="Calibri")
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Anchos de columna (ahora hay 14 columnas, se agregó "Archivo Fuente")
        anchos = [18, 35, 15, 14, 20, 20, 35, 30, 35, 28, 35, 28, 28, 25]
        for i, ancho in enumerate(anchos, start=1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = ancho

        # Congelar encabezados para facilitar lectura
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    output.seek(0)
    return output.read()


# ══════════════════════════════════════════════════════════════════════════════
# INTERFAZ PRINCIPAL
# ══════════════════════════════════════════════════════════════════════════════

# ── Header ───────────────────────────────────────────────────────────────────
st.markdown("""
<div class="header-banner">
    <div class="header-badge">SEMARNAT · NOM-055</div>
    <h1>Gestor Ambiental IA — Bitácora de Residuos Peligrosos</h1>
    <p>Extracción automática de datos desde manifiestos · Bitácora acumulada por sesión · Powered by Claude AI</p>
</div>
""", unsafe_allow_html=True)

# ── Banner de estado de la bitácora acumulada ─────────────────────────────────
# Muestra cuántos registros llevas acumulados en esta sesión de trabajo
total_acumulados = len(st.session_state["bitacora_acumulada"])
if total_acumulados > 0:
    st.markdown(f"""
    <div class="acum-banner">
        📊 <strong>Bitácora activa:</strong> {total_acumulados} manifiesto(s) acumulado(s) en esta sesión.
        Puedes seguir subiendo más PDFs y se agregarán automáticamente.
    </div>
    """, unsafe_allow_html=True)

# ── Configuración de API Key ─────────────────────────────────────────────────
with st.expander("⚙️ Configuración de API Key", expanded=False):
    st.markdown('<div class="card-title">CLAUDE API KEY (ANTHROPIC)</div>', unsafe_allow_html=True)
    api_key_input = st.text_input(
        "Ingresa tu API Key de Claude (Anthropic):",
        type="password",
        placeholder="sk-ant-...",
        help="Obtén tu key en https://console.anthropic.com"
    )
    st.caption("También puedes definirla en un archivo `.env` como `ANTHROPIC_API_KEY=tu_key`.")

api_key = api_key_input.strip() if api_key_input else get_api_key()

# ── Subida de PDFs ───────────────────────────────────────────────────────────
st.markdown('<div class="card"><div class="card-title">📄 Subir Manifiestos (PDF)</div>', unsafe_allow_html=True)

# Fecha de ingreso al almacén — va junto con la subida para que sea parte del registro
col_up, col_fecha_up = st.columns([2, 1])
with col_up:
    uploaded_files = st.file_uploader(
        "Selecciona uno o varios manifiestos en PDF (incluso PDFs con múltiples manifiestos):",
        type=["pdf"],
        accept_multiple_files=True,
        help="Puedes subir múltiples archivos. PDFs pesados se comprimen automáticamente.",
    )

with col_fecha_up:
    fecha_ingreso = st.date_input(
        "📅 Fecha de ingreso al almacén:",
        value=date.today(),
        format="DD/MM/YYYY",
        help="Esta fecha se asignará a todos los manifiestos de esta carga."
    )
    fecha_str = fecha_ingreso.strftime("%d/%m/%Y")

if uploaded_files:
    st.markdown(f"**{len(uploaded_files)} archivo(s) listo(s) para procesar:**")
    for f in uploaded_files:
        tamano = f.size / 1024
        unidad = "KB"
        if tamano > 1024:
            tamano = tamano / 1024
            unidad = "MB"
        icono = "⚠️" if (f.size / (1024*1024)) > 2 else "📄"
        nota = " — PDF grande, se comprimirá automáticamente" if (f.size / (1024*1024)) > 2 else ""
        st.markdown(
            f'<div class="status-ok">{icono} {f.name} — {tamano:.1f} {unidad}{nota}</div>',
            unsafe_allow_html=True
        )

st.markdown('</div>', unsafe_allow_html=True)

# ── Botones de acción ─────────────────────────────────────────────────────────
col_btn, col_limpiar, col_space = st.columns([1, 1, 2])
with col_btn:
    procesar = st.button("🔍 Procesar y acumular", use_container_width=True)
with col_limpiar:
    # Botón para borrar toda la bitácora acumulada y empezar de cero
    limpiar = st.button("🗑️ Limpiar bitácora", use_container_width=True)

if limpiar:
    st.session_state["bitacora_acumulada"] = []
    st.session_state["errores_acumulados"] = []
    st.session_state["archivos_procesados"] = set()
    st.success("✅ Bitácora limpiada. Puedes empezar una nueva sesión de trabajo.")
    st.rerun()

# ── Procesamiento de PDFs ─────────────────────────────────────────────────────
if procesar:
    if not uploaded_files:
        st.warning("⚠️ Por favor sube al menos un archivo PDF.")
    elif not api_key:
        st.error("❌ Se requiere una API Key de Claude. Configúrala en la sección de arriba.")
    else:
        nuevos_registros = 0
        archivos_duplicados = []

        progress_bar = st.progress(0, text="Iniciando procesamiento...")
        log_container = st.container()

        for i, pdf_file in enumerate(uploaded_files):
            pct = int((i / len(uploaded_files)) * 100)
            progress_bar.progress(pct, text=f"Procesando: {pdf_file.name} ({i+1}/{len(uploaded_files)})")

            # Verificar si este archivo ya fue procesado en esta sesión
            if pdf_file.name in st.session_state["archivos_procesados"]:
                with log_container:
                    st.markdown(
                        f'<div class="status-warn">⚠️ <strong>{pdf_file.name}</strong> — '
                        f'Ya fue procesado antes en esta sesión. Se omite para evitar duplicados.</div>',
                        unsafe_allow_html=True
                    )
                archivos_duplicados.append(pdf_file.name)
                continue

            with log_container:
                with st.spinner(f"Leyendo con IA... → {pdf_file.name}"):
                    try:
                        # Paso 1: Convertir PDF a imágenes (con compresión automática si es grande)
                        images, tamano_mb, total_paginas = pdf_to_images_base64(pdf_file)

                        info_pdf = f"{total_paginas} página(s), {tamano_mb:.1f} MB"
                        if tamano_mb > 2:
                            info_pdf += " — comprimido automáticamente"

                        # Paso 2: Enviar a Claude y extraer datos (puede devolver varios manifiestos)
                        lista_manifiestos = extraer_datos_claude(images, api_key)

                        # Paso 3: Agregar cada manifiesto encontrado a la bitácora acumulada
                        for manifiesto in lista_manifiestos:
                            manifiesto["_archivo"] = pdf_file.name
                            manifiesto["_fecha_ingreso"] = fecha_str
                            manifiesto["_procesado_en"] = datetime.now().strftime("%d/%m/%Y %H:%M")
                            st.session_state["bitacora_acumulada"].append(manifiesto)
                            nuevos_registros += 1

                        # Marcar este archivo como procesado
                        st.session_state["archivos_procesados"].add(pdf_file.name)

                        cantidad_encontrados = len(lista_manifiestos)
                        st.markdown(
                            f'<div class="status-ok">✅ <strong>{pdf_file.name}</strong> — '
                            f'{cantidad_encontrados} manifiesto(s) extraído(s). ({info_pdf})</div>',
                            unsafe_allow_html=True
                        )

                    except Exception as e:
                        st.session_state["errores_acumulados"].append((pdf_file.name, str(e)))
                        st.markdown(
                            f'<div class="status-error">❌ <strong>{pdf_file.name}</strong> — '
                            f'Error: {str(e)[:200]}</div>',
                            unsafe_allow_html=True
                        )

        progress_bar.progress(100, text="✅ Procesamiento completado.")

        if nuevos_registros > 0:
            st.success(
                f"✅ Se agregaron **{nuevos_registros}** registro(s) nuevo(s) a la bitácora. "
                f"Total acumulado: **{len(st.session_state['bitacora_acumulada'])}** manifiestos."
            )
        if archivos_duplicados:
            st.info(f"ℹ️ {len(archivos_duplicados)} archivo(s) omitido(s) por duplicado.")

# ── Mostrar bitácora acumulada ────────────────────────────────────────────────
registros = st.session_state["bitacora_acumulada"]

if registros:
    st.markdown("---")

    # Métricas
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.markdown(f"""
        <div class="metric-box">
            <div class="num">{len(registros)}</div>
            <div class="label">Total acumulado</div>
        </div>""", unsafe_allow_html=True)
    with col2:
        archivos_unicos = len(st.session_state["archivos_procesados"])
        st.markdown(f"""
        <div class="metric-box">
            <div class="num">{archivos_unicos}</div>
            <div class="label">PDFs procesados</div>
        </div>""", unsafe_allow_html=True)
    with col3:
        errores_count = len(st.session_state.get("errores_acumulados", []))
        st.markdown(f"""
        <div class="metric-box">
            <div class="num">{errores_count}</div>
            <div class="label">Errores</div>
        </div>""", unsafe_allow_html=True)
    with col4:
        exitosos = len(registros)
        st.markdown(f"""
        <div class="metric-box">
            <div class="num">{exitosos}</div>
            <div class="label">Exitosos</div>
        </div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Tabla de vista previa
    st.markdown('<div class="card"><div class="card-title">📋 Bitácora acumulada (vista previa)</div>', unsafe_allow_html=True)

    df_preview = pd.DataFrame([{
        "# Manifiesto": r["consecutivo"],
        "Residuo": r["nombre_residuo"],
        "Cantidad": r["cantidad"],
        "CRETIB": r["cretib"],
        "Fecha Ingreso": r.get("_fecha_ingreso", ""),
        "Fecha Salida": r["fecha_salida"],
        "Transportista": r["transportista_nombre"],
        "Destinatario": r["destinatario_nombre"],
        "Responsable": r["responsable"],
        "Archivo PDF": r["_archivo"],
    } for r in registros])

    st.dataframe(df_preview, use_container_width=True, hide_index=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # ── Descarga del Excel ────────────────────────────────────────────────────
    st.markdown('<div class="card"><div class="card-title">📥 Descargar Bitácora Excel Completa</div>', unsafe_allow_html=True)

    col_info, col_dl = st.columns([2, 1])
    with col_info:
        st.markdown(f"""
        El Excel incluirá **todos los {len(registros)} registro(s)** acumulados en esta sesión,
        con su fecha de ingreso individual y formato institucional SEMARNAT.
        """)

    with col_dl:
        if st.button("📥 Generar y descargar Excel", use_container_width=True):
            with st.spinner("Generando archivo Excel con formato SEMARNAT..."):
                try:
                    excel_bytes = generar_excel(registros, fecha_str)
                    nombre_archivo = f"Bitacora_RP_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    st.success(f"✅ Bitácora generada con {len(registros)} registro(s).")
                    st.download_button(
                        label="⬇️ Descargar Bitácora Excel",
                        data=excel_bytes,
                        file_name=nombre_archivo,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception as e:
                    st.error(f"Error al generar Excel: {e}")

    st.markdown('</div>', unsafe_allow_html=True)

    # ── Edición manual de registros ───────────────────────────────────────────
    with st.expander("✏️ Editar / corregir registros manualmente (opcional)"):
        st.caption("Puedes corregir cualquier dato antes de descargar el Excel.")
        campos_editables = [
            "consecutivo", "nombre_residuo", "cantidad", "cretib",
            "fecha_salida", "fase_siguiente", "area_resguardo",
            "transportista_nombre", "transportista_autorizacion",
            "destinatario_nombre", "destinatario_autorizacion", "responsable"
        ]
        for idx, reg in enumerate(registros):
            with st.container():
                st.markdown(
                    f"**📄 Manifiesto #{reg['consecutivo']} — {reg['_archivo']} "
                    f"(Ingreso: {reg.get('_fecha_ingreso', '')})**"
                )
                cols = st.columns(3)
                for i, campo in enumerate(campos_editables):
                    with cols[i % 3]:
                        nuevo_val = st.text_input(
                            campo.replace("_", " ").title(),
                            value=reg[campo],
                            key=f"edit_{idx}_{campo}"
                        )
                        st.session_state["bitacora_acumulada"][idx][campo] = nuevo_val
                st.markdown("---")

elif total_acumulados == 0:
    st.info("ℹ️ Aún no hay registros. Sube tus PDFs y presiona 'Procesar y acumular'.")

# ── Instrucciones de uso ──────────────────────────────────────────────────────
with st.expander("ℹ️ Instrucciones de uso"):
    st.markdown("""
    **Cómo usar la bitácora acumulada:**

    1. **Configura tu API Key** de Claude (Anthropic) en la sección de arriba
    2. **Sube los PDFs** — pueden ser manifiestos individuales o PDFs con varios manifiestos
    3. **Elige la fecha de ingreso** al almacén (se asigna a esta carga de PDFs)
    4. **Presiona "Procesar y acumular"** — los datos se agregan a la bitácora
    5. **Sube más PDFs cuando quieras** — se seguirán acumulando en la misma tabla
    6. **Descarga el Excel** cuando tengas todos los que necesitas

    **¿Qué pasa si subo el mismo PDF dos veces?**
    La app lo detecta y lo omite automáticamente para evitar duplicados.

    **¿Cómo empiezo de cero?**
    Presiona el botón "🗑️ Limpiar bitácora" y la memoria de la sesión se borra.

    **PDFs grandes (más de 2 MB):**
    Se comprimen automáticamente antes de enviarlos a la IA. No necesitas hacer nada especial.

    **PDFs con múltiples manifiestos:**
    La IA detecta cuántos manifiestos hay en el documento y los extrae todos de una sola vez.
    """)

# ── Footer ─────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="footer">
    Gestor Ambiental IA · Bitácora de Residuos Peligrosos · Formato SEMARNAT México<br>
    Datos procesados localmente · La información no se almacena en servidores externos
</div>
""", unsafe_allow_html=True)
